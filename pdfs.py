from datetime import datetime
import os
import re
from re import X
import shutil
import PyPDF2 
import pandas as pd    
import camelot
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from pdf2image import convert_from_path
from PIL import Image
from multiprocessing import Process,Queue
words=['recycled','DEVELOPMENT NOTE','nylon','elastane','polyester','spandex']
startTime = datetime.now()
print(datetime.now() - startTime)
Q=Queue()
Q1=Queue()
df3=pd.DataFrame()
df4=pd.DataFrame()
df1=pd.DataFrame()
direc=os.getcwd()
path2=direc+'/pdf scraper fields v2.xlsx'
wb=openpyxl.load_workbook(path2)
sheet=wb.worksheets[0]
sheet2=wb.worksheets[1]
sheet3=wb.worksheets[2]

#print(sheet, sheet2, sheet3)

lis3_rw=3
rw=4
col=4
rw_fail=4
junn=0
color_lis=[]
jun,jun1,jun2,jun3=[],[],[],[]
lads,lads1,lads2=[],[],[]


def table1(path1,file1):
    try:
        tables1 = camelot.read_pdf(os.path.join(path1, file1),flavor='stream',pages='1',table_areas=['1,534,202,465','1,448,209,120','560,540,840,20','694,555,788,535', '1,131,173,15'], columns=['67,105,144','','','',''], edge_tol=150,split_text=True)
#        camelot.plot(tables1[4], kind='contour').show()
#        plt.show(block=True)
        Q1.put(tables1)
    except Exception as e:
        Q1.put(e)
def table(path1,file1):
    try:
        tables = camelot.read_pdf(os.path.join(path1, file1),flavor='stream',pages='1',table_areas=['1,534,202,465','1,448,209,120','560,540,840,115','694,555,788,535', '1,131,173,15'],columns=['67,105,144','','','',''],edge_tol=150,split_text=True)
        # camelot.plot(tables1[3], kind='grid').show()
        # plt.show()
        Q.put(tables)
    except Exception as e:
        Q.put(e)

'''testing
def table1(path1,file1):
    try:
        tables1 = camelot.read_pdf(os.path.join(path1, file1),flavor='stream',pages='1',table_areas=['1,534,202,465','1,448,209,120','560,540,840,20','694,555,788,535'],edge_tol=150)
        
        return tables1
    except Exception as e:
        print("tabl1", e)
        return e
        
        
def table(path1,file1):
    try:
        tables = camelot.read_pdf(os.path.join(path1, file1),flavor='stream',pages='1',table_areas=['1,534,202,465','1,448,209,120','560,540,840,115','694,555,788,535'],columns=['67,105,144','','',''],edge_tol=150,split_text=True)
        
        return tables
    except Exception as e:
        print("table", e)
        return e
'''

def collar_trim(df):
    collar=[]
    trim=[]
    collar2=[]
    trim2=[]
    col_l2=[]
    col_l3=[]
    trm_l2=[]
    trm_l3=[]
    for index,rows in df.iterrows():
        for i in range(0,len(df.columns)):
            co=len(collar)
            tr=len(trim)
            if 'collar' in str(rows[i]).lower():
                collar.append(str(rows[i]).replace('COLLARS & TRIMS','').replace('COLLAR: ','').replace('COLLAR','').replace(',,',''))
                try:
                    if (len(collar)>co )and (str(collar[((len(collar))-1)]).startswith('D') or  str(collar[((len(collar))-1)]).startswith('P')  or re.search(r'^\d',str(collar[((len(collar))-1)]))):
                        if ('COLLAR:' not in str(df.iloc[index+1][i])) and ('TRIM:' not in str(df.iloc[index+1][i])):
                            col_l2.append(str(df.iloc[index+1][i]).replace('INFORMATION REQUIRED',''))
                except:
                    pass
                try:
                    if (len(collar)>co )and (str(collar[((len(collar))-1)]).startswith('D') or  str(collar[((len(collar))-1)]).startswith('P')  or re.search(r'^\d',str(collar[((len(collar))-1)]))):
                        if ('COLLAR:' not in str(df.iloc[index+2][i])) and ('TRIM:'not in str(df.iloc[index+2][i])):
                            col_l3.append(str(df.iloc[index+2][i]).replace('INFORMATION REQUIRED',''))
                except:
                    pass
            if 'trim' in str(rows[i]).lower():
                
                trim.append(str(rows[i]).replace('COLLARS & TRIMS','').replace('TRIM: ','').replace('TRIM','').replace(',,',''))
                try:
                    if len(trim)>tr and (str(trim[((len(trim))-1)]).startswith('D') or  str(trim[((len(trim))-1)]).startswith('P')  or re.search(r'^\d',str(trim[((len(trim))-1)]))):
                        if 'COLLAR:' not in str(df.iloc[index+1][i]) and 'TRIM:' not in str(df.iloc[index+1][i]):
                            trm_l2.append(str(df.iloc[index+1][i]).replace('INFORMATION REQUIRED',''))
                except:
                    pass
                try:
                    if len(trim)>tr and (str(trim[((len(trim))-1)]).startswith('D') or  str(trim[((len(trim))-1)]).startswith('P')  or re.search(r'^\d',str(trim[((len(trim))-1)]))):
                        if 'COLLAR:' not in str(df.iloc[index+2][i]) and 'TRIM:' not in str(df.iloc[index+2][i]):
                            trm_l3.append(str(df.iloc[index+2][i]).replace('INFORMATION REQUIRED',''))
                except:
                    pass
    # collar=(','.join(collar).replace('COLLARS & TRIMS','').replace('COLLAR: ','').replace('COLLAR','').replace(',,',''))
    # trim=(','.join(trim).replace('COLLARS & TRIMS','').replace('TRIM: ','').replace('TRIM','').replace(',,',''))
   
    for i in collar:
        
        if str(i).startswith('D')or  str(i).startswith('P')  or re.search(r'^\d',str(i)):
            collar2.append(i)
        else:
            pass
    
    for i in trim:
        if str(i).startswith('D') or str(i).startswith('P') or re.search(r'^\d',str(i)):
            trim2.append(i)
        else:
            pass
    collar=','.join(collar2)
    trim=','.join(trim2)
    # collar=','.join(collar)
    # trim=(','.join(trim))
    # form1=r"[DS]\d+"
    # form2=r"[DST]\d+"
    # collar=re.findall(form1,collar)
    return[collar,trim,collar2,trim2,col_l2,col_l3,trm_l2,trm_l3]
# color code and description function
def color_code(df):
    x,x1=[],[]
    y,y1=[],[]
    z,z1=[],[]
    sub=[]
    c1=[]
    c11=0
    c2=0
    c22=0
    c3=0
    c33=0

    lis=[]
    lis2=[]
    lis1=[]
    lis3=[]
    color_frmt = re.compile(r"[C]\-[0-9]{3}")
    color_frmt2=r"[E]\d{3,5}"
    clr_cnt=0
    
    if (len(df.columns))<=2:
        
        for i in range(0,(len(df.index))):  
            if 'gsm' in df.iloc[i,0]:
                clr_cnt=clr_cnt+1
            if len(df.columns)>1:
                if ('Recycled' in df.iloc[i,0])  or ('Chamois Leather' in df.iloc[i,0]) or ('DEVELOPMENT NOTE' in df.iloc[i,0])  or ('nylon' in str(df.iloc[i,0]).lower() )or ('elastane' in str(df.iloc[i,0]).lower())or ('polyester' in str(df.iloc[i,0]).lower() )or ('spandex' in str(df.iloc[i,0]).lower() ) :
                        lis.append(i)
                        c1.append(0)
                elif ('Recycled' in df.iloc[i,1])  or ('Chamois Leather' in df.iloc[i,1]) or ('DEVELOPMENT NOTE' in df.iloc[i,1]) or ('nylon' in str(df.iloc[i,1]).lower() )or ('elastane' in str(df.iloc[i,1]).lower())or ('polyester' in str(df.iloc[i,1]).lower() )or ('spandex' in str(df.iloc[i,1]).lower() ) :
                        lis.append(i)
                        c1.append(1) 
            else:
                if ('Recycled' in df.iloc[i,0])  or ('Chamois Leather' in df.iloc[i,0]) or ('DEVELOPMENT NOTE' in df.iloc[i,0]) or ('nylon' in str(df.iloc[i,0]).lower() )or ('elastane' in str(df.iloc[i,0]).lower())or ('polyester' in str(df.iloc[i,0]).lower() )or ('spandex' in str(df.iloc[i,0]).lower() ) :
                    lis.append(i)
                    c1.append(0) 
        print(lis) 
        print(c1)                            
        # or ('composition' in str(df.iloc[i,0]).lower() )
        if int(lis[0])<2:
            lis.pop(0)
            c1.pop(0)
        
        if len(lis)>=1:
            if c1[0]==0:
                try:
                    if (len (re.findall(color_frmt2,str(df.iloc[int(lis[0])-2,0])))) >0:
                        
                        lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),0])]
                except:
                    pass
                try:
                    if (len (re.findall(color_frmt2,str(df.iloc[int(lis[0])-3,0])))) >0:
                        lis1=[str(df.iloc[int(lis[0])-3,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),0])]
                except:
            
                    pass
                try:
                    if ((len (re.findall(color_frmt2,str(df.iloc[int(lis[0])-2,0])))) ==0) and ((len (re.findall(color_frmt2,str(df.iloc[int(lis[0])-3,0])))) ==0):
                        lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),0])]
                except:
            
                    pass
            else:
                lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            
            if 'SUBLIMATED' in df.iloc[int(lis[0])+1,0]:
                sub.append('Yes')
            else:
                sub.append('No')  
        try:
            
            if (len (re.findall(color_frmt,str(df.iloc[(int(lis[0]))+2,0])))) >0:
                
                if len(lis)>1 :
                    
                    for k in range(((int(lis[0]))+1),((int(lis[1]))-2),2):
                        x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
                else:
                    if (((len(df.index))-1) - (int(lis[0])))>3:
                        for k in range(((int(lis[0]))+1),(len(df.index)),2):
                            x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
                        else:
                            for k in range(((int(lis[0]))+1),(len(df.index))-1,2):
                                x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
        
        except:
            pass
        
        if len(lis)>=2 :
            if ((len(df.index))-1)> lis[1]:
                if 'SUBLIMATED' in df.iloc[int(lis[1])+1,0]:
                    sub.append('Yes')
                else:
                    sub.append('No')
            
                if c1[1]==0:
                    lis2=[str(df.iloc[int(lis[1])-2,0]),str(df.iloc[int(lis[1])-1,0]),str(df.iloc[int(lis[1]),0])]
                    
                else:
                    lis2=[str(df.iloc[int(lis[1])-2,0]),str(df.iloc[int(lis[1])-1,0]),str(df.iloc[int(lis[1]),1])]
            
        if len(lis)>=3 :
            if 'SUBLIMATED' in df.iloc[int(lis[2])+1,0]:
                sub.append('Yes')
            else:
                sub.append('No')
            if c1[2]==0:
                lis3=[str(df.iloc[int(lis[2])-2,0]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),0])]
            else:
                lis3=[str(df.iloc[int(lis[2])-2,0]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),1])]
           
    else:
        
        
        # color_frmt2 = re.compile(r"[E]\[0-9]{1}")  
        
        
        for i in range(0,(len(df.index))):  
            if 'gsm' in df.iloc[i,0]:
                clr_cnt=clr_cnt+1
            if ('Recycled' in df.iloc[i,1]) or ('Chamois Leather' in df.iloc[i,1]) or ('DEVELOPMENT NOTE' in df.iloc[i,1]) or ('nylon' in str(df.iloc[i,1]).lower() )or ('elastane' in str(df.iloc[i,1]).lower()) or ('polyester' in str(df.iloc[i,1]).lower() )or ('spandex' in str(df.iloc[i,1]).lower()) or ('cotton drill' in str(df.iloc[i,1]).lower()):
                lis.append(i)
                c1.append(1)
            elif ('Recycled' in df.iloc[i,0]) or ('Chamois Leather' in df.iloc[i,0]) or ('DEVELOPMENT NOTE' in df.iloc[i,0]) or ('nylon' in str(df.iloc[i,0]).lower() )or ('elastane' in str(df.iloc[i,0]).lower()) or ('polyester' in str(df.iloc[i,0]).lower() )or ('spandex' in str(df.iloc[i,0]).lower()) or ('cotton drill' in str(df.iloc[i,0]).lower()):
                lis.append(i)
                c1.append(0)
        
         
        if int(lis[0])<2:
            lis.pop(0)
            c1.pop(0)
        print(lis)
        print(c1)
        if len(lis)>=1:
            if str(df.iloc[int(lis[0])-2,1]) !='':
                lis1=[str(df.iloc[int(lis[0])-2,1]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            
            # elif str(df.iloc[int(lis[0])-2,0]) !='' and str(df.iloc[int(lis[0])-2,0]) !='.' :
            #     lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            # else:
            #     lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            elif len (re.findall(color_frmt2, df.iloc[(int(lis[0]))-2,1]))>0:
            #     print('in')
                lis1=[str(df.iloc[int(lis[0])-2,1]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            elif len (re.findall(color_frmt2, str(df.iloc[(int(lis[0]))-3,1])))>0:
            #     print('in')
                lis1=[str(df.iloc[int(lis[0])-3,1]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            elif len (re.findall(color_frmt2, df.iloc[(int(lis[0]))-2,0]))>0:
                lis1=[str(df.iloc[int(lis[0])-2,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
            # print(lis1)
            else:
                try:
                    if str(df.iloc[int(lis[0])-2,1]) =='' and ('E0' in str(df.iloc[int(lis[0])-3,0])) :
                        lis1=[str(df.iloc[int(lis[0])-3,0]),str(df.iloc[int(lis[0])-1,0]),str(df.iloc[int(lis[0]),1])]
                except:
                    pass

            if 'SUBLIMATED' in df.iloc[int(lis[0])+1,0]:
                sub.append('Yes')
            elif 'SUBLIMATED' in df.iloc[int(lis[0])+1,(len(df.columns)-1)]:
                sub.append('Yes')
            else:
                sub.append('No')
            try:
                if (len (re.findall(color_frmt,str(df.iloc[(int(lis[0]))+2,0])))) >0:
                    if len(lis)>1:
                        for k in range(((int(lis[0]))+1),((int(lis[1]))-2),2):
                            x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
                    else:
                        
                        if ((len(df.index)-1) - (int(lis[0])))>3:
                            
                            for k in range(((int(lis[0]))+1),(len(df.index)),2):
                                x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
                        else:
                            
                            for k in range(((int(lis[0]))+1),(len(df.index)-1),2):
                                x=x+[str(df.iloc[k,u]+':'+df.iloc[k+1,u]) for u in range(0,len(df.columns))]
            except:
                pass
        
        if len(lis)>=2:
            step=0
            if 'SUBLIMATED' in df.iloc[int(lis[1])+1,0]:
                sub.append('Yes')
            elif 'SUBLIMATED' in df.iloc[int(lis[1])+1,(len(df.columns)-1)]:
                sub.append('Yes')
            else:
                sub.append('No')
            if c1[1]==1:
                lis2=[str(df.iloc[(int(lis[1]))-2,1]),str(df.iloc[(int(lis[1]))-1,0]),str(df.iloc[(int(lis[1])),1])]
            else:
                lis2=[str(df.iloc[(int(lis[1]))-2,1]),str(df.iloc[(int(lis[1]))-1,0]),str(df.iloc[(int(lis[1])),0])]
            print(len(df.index))
            try:
                
                if (int(lis[1]))+2 >= ((len(df.index))+1):
                    
                    if len(re.findall(color_frmt,df.iloc[(int(lis[1]))+2,0])) >0 and len(lis)==2:
                        
                        for l in range(((int(lis[1]))+1),((int(len(df.index)))-1),2):
                            y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
                    elif len(re.findall(color_frmt,df.iloc[(int(lis[1]))+2,1])) >0 and len(lis)==2:
                        
                        for l in range(((int(lis[1]))+1),((int(len(df.index)))-1),2):
                            y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
                    else:
                        
                        for l in range(((int(lis[0]))+1),((int(lis[1]))-2),2):
                                y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
                    step=1
            except:
                pass
            try:        
                if len(lis)>2 and len(re.findall(color_frmt,df.iloc[(int(lis[1]))+2,0])) ==0:
                    step=1
                    for l in range(((int(lis[1]))+1),((int(lis[2]))-2),2):
                        y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
            except:
                pass    
            try:
                if len(lis)==2 and len(re.findall(color_frmt,df.iloc[(int(lis[1]))+2,1])) >0:
                    step=1
                    for l in range(((int(lis[1]))+1),((int(len(df.index)))-1),2):
                        y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
            except:
                pass
            try:    
                if step==0:
                                    
                    for l in range(((int(lis[1]))+1),((int(lis[2]))-2),2):
                        y=y+[str(df.iloc[l,u]+':'+df.iloc[l+1,u]) for u in range(0,len(df.columns))]
            except:
                pass

        
        if len(lis)>3 :
            if 'SUBLIMATED' in df.iloc[int(lis[2])+1,0]:
                sub.append('Yes')
            elif 'SUBLIMATED' in df.iloc[int(lis[2])+1,(len(df.columns)-1)]:
                sub.append('Yes')
            else:
                sub.append('No')
            if c1[3]==1:
                lis3=[str(df.iloc[int(lis[2])-2,1]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),1])]
            else:
                lis3=[str(df.iloc[int(lis[2])-2,1]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),0])]
            for m in range(((int(lis[2]))+1),((int(lis[3]))-2),2):
                z=z+[str(df.iloc[m,u]+':'+df.iloc[m+1,u]) for u in range(0,len(df.columns))]
        
        if (len(lis)==3) and ('DEVELOPMENT NOTE' not in str(df.iloc[int(lis[2]),1])) and ('DEVELOPMENT NOTE' not in str(df.iloc[int(lis[2]),0])) :
            
            if 'SUBLIMATED' in df.iloc[int(lis[2])+1,0]:
                sub.append('Yes')
            elif 'SUBLIMATED' in df.iloc[int(lis[2])+1,(len(df.columns)-1)]:
                sub.append('Yes')
            else:
                sub.append('No')
            if c1[2]==1:
                lis3=[str(df.iloc[int(lis[2])-2,1]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),1])]
            else:
                lis3=[str(df.iloc[int(lis[2])-2,1]),str(df.iloc[int(lis[2])-1,0]),str(df.iloc[int(lis[2]),0])]
            for m in range(((int(lis[2]))+1),(int(len(df.index))),2):
                try:
                    z=z+[str(df.iloc[m,u]+':'+str(df.iloc[m+1,u])) for u in range(0,len(df.columns))]
                except:
                    pass
                   
    # x=','.join(x)
    # y=','.join(y)
    # z=','.join(z)
    # print(x)
    # print(y )
    # print(z)
    print(lis1)
    print(lis2)
    print(lis3)
    if len(x)>1:
        # d.append(re.split(',)
        x1=x
        x=','.join(re.findall(color_frmt, ','.join(x)))

    else: 
        x1=x
        x=','.join(x)
    if len(y)>1:
        y1=y
        y=','.join(re.findall(color_frmt, ','.join(y)))
    else:
        y1=y
        y=','.join(y)
    if len(z)>1:
        z1=z
        z=','.join(re.findall(color_frmt, ','.join(z)))
    else:
        z1=z
        z=','.join(z)
    
    return [x,y,z,sub,lis1,lis2,lis3,x1,y1,z1]
## Image cropping function
def ImageCrop(img):
    img_nme=img
    img = Image.open(img)
    left = 725
    top = 200
    right = 2425
    bottom = 2000
    img_res = img.crop((left, top, right, bottom))
    with open(img_nme, 'w') as f:
        img_res.save(img_nme,'PNG')

print("Reading directory ------> ", direc)

isExist1 = os.path.exists(direc+'/failed/')
isExist2 = os.path.exists(direc+'/Image/')
if not isExist1:
    os.makedirs(direc+'/failed/')
if not isExist2:
    os.makedirs(direc+'/Image/')
isExist1 =direc+'/failed/'
isExist2 =direc+'/Image/'

for path, subdirs, files in os.walk(direc):
    if str(path.replace(direc,'')) == '/failed':
        pass
    else:
        
        for file in files:

            if file.lower().endswith('.pdf'):
                
                print("\nGot PDF: ", file)
                
                folder_name=os.path.basename(path)
                
                try:

                    p1 = Process(target=table, args=(path,file))
                    p2 = Process(target=table1, args=(path,file))
                    p1.start()
                    p2.start()
                    tables=(Q.get())
                    tables1=Q1.get()
                    p1.join()
                    p2.join()
                        
                    #tables = table(path, file)
                    #tables1 = table1(path, file)
                    
                    #print("tables----->", tables, tables1)
                    
                    try:
                        (tables1[1].df).to_excel((direc+'/tables/'+'first1'+file+'.xlsx'),index=False)
                        
                        print("tables1[1].df--- ", tables1[1].df)
                        
                        sheet.cell(row=rw,column=4).value=rw-3
                        sheet.cell(row=rw,column=5).value=folder_name
                        sheet.cell(row=rw,column=30).value=file
                        sheet.cell(row=rw,column=1).value=(str(file)).split(' ')[0]
                        sheet.cell(row=rw,column=2).value=(str(tables[0].df.iloc[0,3])).replace('DATE: ','')
                        sheet.cell(row=rw,column=3).value=(str(file)).replace('.pdf','').replace('.PDF','')
                        
                        # (tables[1].df).to_excel(('first5'+'.xlsx'),index=False)
                        if (('mens' in str(tables[1].df.iloc[1,0]).lower()) or('mens' in str(tables[1].df.iloc[1,1]).lower())):

                            
                            man=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('mens' in str(tables[1].df.iloc[1,0]).lower()) or('mens' in str(tables[1].df.iloc[1,1]).lower()))]
                            if 2<len(tables[1].df.columns)<4:
                                sheet.cell(row=rw,column=6).value=man[0]#tables[1].df.iloc[1,0]
                                sheet.cell(row=rw,column=7).value=man[1]#tables[1].df.iloc[1,2]
                                sheet.cell(row=rw,column=8).value=man[2]#
                            elif len(tables[1].df.columns)==2:
                                sheet.cell(row=rw,column=6).value=((str(man[0])).split('\n'))[1]#tables[1].df.iloc[1,0]
                                sheet.cell(row=rw,column=7).value=((str(man[0])).split('\n'))[0]#tables[1].df.iloc[1,2]
                                sheet.cell(row=rw,column=8).value=man[1]

                            else:
                                sheet.cell(row=rw,column=6).value=man[0]+' '+man[1]#tables[1].df.iloc[1,0]
                                sheet.cell(row=rw,column=7).value=man[2]#tables[1].df.iloc[1,2]
                                sheet.cell(row=rw,column=8).value=man[3]
                           
                        if (('ladies' in str(tables[1].df.iloc[1,0]).lower()) or('ladies' in str(tables[1].df.iloc[1,1]).lower())) or (('ladies' in str(tables[1].df.iloc[2,0]).lower()) or('ladies' in str(tables[1].df.iloc[2,1]).lower())):
                            
                            try:
                                lads1=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('ladies' in str(tables[1].df.iloc[1,0]).lower()) or('ladies' in str(tables[1].df.iloc[1,1]).lower()))]
                            except:
                                pass
                            try:
                                lads2=[str(tables[1].df.iloc[2,u]) for u in range (0,len(tables[1].df.columns)) if (('ladies' in str(tables[1].df.iloc[2,0]).lower()) or('ladies' in str(tables[1].df.iloc[2,1]).lower()))]
                            except:
                                pass
                            lads=lads1+lads2
                            
                            if 2< len(tables[1].df.columns)<4:
                                sheet.cell(row=rw,column=9).value=lads[0]#tables[1].df.iloc[1,0]
                                sheet.cell(row=rw,column=10).value=lads[1]#tables[1].df.iloc[1,2]
                                sheet.cell(row=rw,column=11).value=lads[2]#
                            elif len(tables[1].df.columns)==2:
                                try:
                                    sheet.cell(row=rw,column=9).value=(str(lads[0])).replace('LADIES \n','').split('\n')[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=10).value=(str(lads[0])).replace('LADIES \n','').split('\n')[0]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=11).value=lads[1]
                                except:
                                    sheet.cell(row=rw,column=9).value=(str(lads[0])).split(' LADIES ')[0]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=10).value=(str(lads[0])).split(' LADIES ')[1]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=11).value=lads[1]

                            else:
                                sheet.cell(row=rw,column=9).value=lads[0]+' '+lads[1]#tables[1].df.iloc[1,0]
                                sheet.cell(row=rw,column=10).value=lads[2]#tables[1].df.iloc[1,2]
                                sheet.cell(row=rw,column=11).value=lads[3]
                        jun=[]
                        jun1,jun2,jun3=[],[],[]
                        junn=0
                        if (len((tables[1].df).index))>3:
                            
                            if (('juniors' in str(tables[1].df.iloc[1,0]).lower()) or('juniors' in str(tables[1].df.iloc[1,1]).lower())) or (('juniors' in str(tables[1].df.iloc[2,0]).lower()) or('juniors' in str(tables[1].df.iloc[2,1]).lower())) or (('juniors' in str(tables[1].df.iloc[3,0]).lower()) or('juniors' in str(tables[1].df.iloc[3,1]).lower())):
                                junn=1
                                try:
                                    jun1=[str(tables[1].df.iloc[3,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[3,0]).lower()) or('juniors' in str(tables[1].df.iloc[3,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun2=[str(tables[1].df.iloc[2,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[2,0]).lower()) or('juniors' in str(tables[1].df.iloc[2,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun3=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[1,0]).lower()) or('juniors' in str(tables[1].df.iloc[1,1]).lower()))]
                                except:
                                    pass
                                jun=jun1+jun3+jun2
                                
                                if 2<len(tables[1].df.columns)<4:
                                    sheet.cell(row=rw,column=12).value=jun[0]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[1]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[2]#
                                elif len(tables[1].df.columns)==2:
                                    sheet.cell(row=rw,column=12).value=((str(jun[0])).split('\n'))[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=((str(jun[0])).split('\n'))[0]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[1]

                                else:
                                    
                                    sheet.cell(row=rw,column=12).value=jun[0]+' '+jun[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[2]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[3]               

                            elif ((('junior' in str(tables[1].df.iloc[1,0]).lower()) or('junior' in str(tables[1].df.iloc[1,1]).lower())) or (('junior' in str(tables[1].df.iloc[2,0]).lower()) or('junior' in str(tables[1].df.iloc[2,1]).lower())) or (('junior' in str(tables[1].df.iloc[3,0]).lower()) or('junior' in str(tables[1].df.iloc[3,1]).lower()))) and (junn==0):
                                print('here')
                                try:
                                    jun1=[str(tables[1].df.iloc[3,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[3,0]).lower()) or('junior' in str(tables[1].df.iloc[3,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun2=[str(tables[1].df.iloc[2,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[2,0]).lower()) or('junior' in str(tables[1].df.iloc[2,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun3=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[1,0]).lower()) or('junior' in str(tables[1].df.iloc[1,1]).lower()))]
                                except:
                                    pass
                                jun=jun1+jun3+jun2
                                
                                if 2<len(tables[1].df.columns)<4:
                                    sheet.cell(row=rw,column=12).value=jun[0]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[1]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[2]#
                                elif len(tables[1].df.columns)==2:
                                    sheet.cell(row=rw,column=12).value=((str(jun[0])).split('\n'))[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=((str(jun[0])).split('\n'))[0]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[1]

                                else:
                                    
                                    sheet.cell(row=rw,column=12).value=jun[0]+' '+jun[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[2]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[3]  



                            # else:
                            #     sheet.cell(row=rw,column=19).value='No'
                        else:
                            if (('juniors' in str(tables[1].df.iloc[1,0]).lower()) or('juniors' in str(tables[1].df.iloc[1,1]).lower())) or (('juniors' in str(tables[1].df.iloc[2,0]).lower()) or('juniors' in str(tables[1].df.iloc[2,1]).lower())) :
                                junn=1
                                try:
                                    jun1=[str(tables[1].df.iloc[3,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[3,0]).lower()) or('juniors' in str(tables[1].df.iloc[3,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun2=[str(tables[1].df.iloc[2,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[2,0]).lower()) or('juniors' in str(tables[1].df.iloc[2,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun3=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('juniors' in str(tables[1].df.iloc[1,0]).lower()) or('juniors' in str(tables[1].df.iloc[1,1]).lower()))]
                                except:
                                    pass
                                jun=jun1+jun3+jun2
                                
                                if 2<len(tables[1].df.columns)<4:
                                    sheet.cell(row=rw,column=12).value=jun[0]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[1]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[2]#
                                elif len(tables[1].df.columns)==2:
                                    sheet.cell(row=rw,column=12).value=((str(jun[0])).split('\n'))[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=((str(jun[0])).split('\n'))[0]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[1]

                                else:
                                    
                                    sheet.cell(row=rw,column=12).value=jun[0]+' '+jun[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[2]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[3]               

                            elif ((('junior' in str(tables[1].df.iloc[1,0]).lower()) or('junior' in str(tables[1].df.iloc[1,1]).lower())) or (('junior' in str(tables[1].df.iloc[2,0]).lower()) or('junior' in str(tables[1].df.iloc[2,1]).lower()))) and (junn==0):
                                try:
                                    jun1=[str(tables[1].df.iloc[3,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[3,0]).lower()) or('junior' in str(tables[1].df.iloc[3,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun2=[str(tables[1].df.iloc[2,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[2,0]).lower()) or('junior' in str(tables[1].df.iloc[2,1]).lower()))]
                                except:
                                    pass
                                try:

                                    jun3=[str(tables[1].df.iloc[1,u]) for u in range (0,len(tables[1].df.columns)) if (('junior' in str(tables[1].df.iloc[1,0]).lower()) or('junior' in str(tables[1].df.iloc[1,1]).lower()))]
                                except:
                                    pass
                                jun=jun1+jun3+jun2
                                
                                if 2<len(tables[1].df.columns)<4:
                                    sheet.cell(row=rw,column=12).value=jun[0]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[1]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[2]#
                                elif len(tables[1].df.columns)==2:
                                    sheet.cell(row=rw,column=12).value=((str(jun[0])).split('\n'))[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=((str(jun[0])).split('\n'))[0]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[1]

                                else:
                                    
                                    sheet.cell(row=rw,column=12).value=jun[0]+' '+jun[1]#tables[1].df.iloc[1,0]
                                    sheet.cell(row=rw,column=13).value=jun[2]#tables[1].df.iloc[1,2]
                                    sheet.cell(row=rw,column=14).value=jun[3]  



                            # else:
                            #     sheet.cell(row=rw,column=19).value='No'
                            
                    except Exception as e:
                        print("1111----->", e)
                        pass

                    try:
                        (tables1[3].df).to_excel((direc+'/tables/'+'first'+file+'.xlsx'),index=False)
                        col=collar_trim(tables1[3].df)
                        sheet.cell(row=rw,column=35).value=col[0]
                        sheet.cell(row=rw,column=36).value=col[1]
                        
                        # print(col[3])
                        # print(col[6])
                        # print(col[7])
                        if len(col[2])>0:
                            i=lis3_rw
                            for x in col[2]:
                                i=i+1
                                sheet3.cell(row=i,column=17).value=x
                                
                            i=lis3_rw
                            for y in col[4]:
                                i=i+1
                                sheet3.cell(row=i,column=18).value=y
                                
                            i=lis3_rw
                            for t in col[5]:
                                i=i+1
                                sheet3.cell(row=i,column=19).value=t
                        
                        if len(col[3])>0:
                            i=lis3_rw
                            for x in col[3]:
                                i=i+1
                                sheet3.cell(row=i,column=21).value=x
                                
                            i=lis3_rw
                            for y in col[6]:
                                i=i+1
                                sheet3.cell(row=i,column=22).value=y
                                
                            i=lis3_rw
                            for t in col[7]:
                                i=i+1
                                sheet3.cell(row=i,column=23).value=t
                                
                        lis3_rw=i+5
                    except:
                        pass

                    try:
                        (tables1[2].df).to_excel((direc+'/tables/'+'first5'+file+'.xlsx'),index=False)
                        fn=color_code(tables1[2].df)
                        
                        # print(fn[4])
                        color_lis=color_lis+fn[7]+fn[8]+fn[9]
                        
                        try:
                            if ' ' in str((fn[4][0])) and (len(str((fn[4][0])).split()))>2:
                            
                                sheet.cell(row=rw,column=15).value=((str((fn[4][0]))).rsplit(' ',1))[0]
                                sheet.cell(row=rw,column=16).value=((str((fn[4][0]))).rsplit(' ',1))[1]
                            elif ' ' in str((fn[4][0])) and len(str((fn[4][0])).split())==2:
                            
                                sheet.cell(row=rw,column=15).value=((str((fn[4][0]))).split())[0]
                                sheet.cell(row=rw,column=16).value=((str((fn[4][0]))).split())[1]
                            else:
                                sheet.cell(row=rw,column=15).value=(str((fn[4][0])))
                                sheet.cell(row=rw,column=16).value=str((fn[4][0]))
                            sheet.cell(row=rw,column=17).value=str((fn[4][1]))
                            sheet.cell(row=rw,column=18).value=str((fn[4][2]))
                            sheet.cell(row=rw,column=19).value=str((fn[3][0]))
                        except:

                            sheet.cell(row=rw,column=15).value=''
                            sheet.cell(row=rw,column=16).value=''
                            sheet.cell(row=rw,column=17).value=''
                            sheet.cell(row=rw,column=18).value=''
                            sheet.cell(row=rw,column=19).value=''

                        try:
                            if ' ' in str((fn[5][0])):
                                sheet.cell(row=rw,column=20).value=((str((fn[5][0]))).rsplit(' ',1))[0]
                                sheet.cell(row=rw,column=21).value=((str((fn[5][0]))).rsplit(' ',1))[1]
                            else:
                                sheet.cell(row=rw,column=20).value=((str((fn[5][0]))))
                                sheet.cell(row=rw,column=21).value=((str((fn[5][0]))))
                            sheet.cell(row=rw,column=22).value=str((fn[5][1]))
                            sheet.cell(row=rw,column=23).value=str((fn[5][2]))
                            sheet.cell(row=rw,column=24).value=str((fn[3][1]))
                        except:
                            sheet.cell(row=rw,column=20).value=''
                            sheet.cell(row=rw,column=21).value=''
                            sheet.cell(row=rw,column=22).value=''
                            sheet.cell(row=rw,column=23).value=''
                            sheet.cell(row=rw,column=24).value=''
                        try:
                            if ' ' in str((fn[6][0])):
                            
                                sheet.cell(row=rw,column=25).value=((str((fn[6][0]))).rsplit(' ',1))[0]
                                sheet.cell(row=rw,column=26).value=((str((fn[6][0]))).rsplit(' ',1))[1]
                            else:
                                sheet.cell(row=rw,column=25).value=(str((fn[6][0])))
                                sheet.cell(row=rw,column=26).value=str((fn[6][0]))
                            sheet.cell(row=rw,column=27).value=str((fn[6][1]))
                            sheet.cell(row=rw,column=28).value=str((fn[6][2]))
                            sheet.cell(row=rw,column=29).value=str((fn[3][1]))
                        except:
                            sheet.cell(row=rw,column=25).value=''
                            sheet.cell(row=rw,column=26).value=''
                            sheet.cell(row=rw,column=27).value=''
                            sheet.cell(row=rw,column=28).value=''
                            sheet.cell(row=rw,column=29).value=''
    
                        try:
                            sheet.cell(row=rw,column=32).value=str(fn[0])
                        except:
                            sheet.cell(row=rw,column=32).value=''
                        try:
                            sheet.cell(row=rw,column=33).value=str(fn[1])
                        except:
                            sheet.cell(row=rw,column=33).value=''
                        try:
                            sheet.cell(row=rw,column=34).value=str(fn[2])
                        except:
                            sheet.cell(row=rw,column=34).value==''
                    except:
                        pass

                    if sheet.cell(row=rw,column=19).value=='Yes' and ((sheet.cell(row=rw,column=32).value)==''):
                        
                        (sheet.cell(row=rw,column=32).value)='SUBLIMATED'
                    if str(sheet.cell(row=rw,column=24).value)=='Yes' and sheet.cell(row=rw,column=33).value=='':
                        sheet.cell(row=rw,column=33).value='SUBLIMATED'
                    if str(sheet.cell(row=rw,column=29).value)=='Yes' and sheet.cell(row=rw,column=34).value=='':
                        sheet.cell(row=rw,column=34).value='SUBLIMATED'
                    if sheet.cell(row=rw-1,column=6).value==None and sheet.cell(row=rw-1,column=7).value==None and sheet.cell(row=rw-1,column=8).value==None and sheet.cell(row=rw-1,column=9).value==None and sheet.cell(row=rw-1,column=10).value==None and sheet.cell(row=rw-1,column=11).value==None and sheet.cell(row=rw-1,column=12).value==None and sheet.cell(row=rw-1,column=13).value==None and sheet.cell(row=rw-1,column=14).value==None:
                        sheet.delete_rows(rw-1)
                        print('move1   ',rw)
                        shutil.copy(path+'/'+file, isExist1)
                        sheet2.cell(row=rw_fail,column=1).value=file     
                        sheet2.cell(row=rw_fail,column=2).value=path.path.replace(direc,'') 
                        rw_fail+=1
                        rw=rw-1
                    else:
                        name=(str(file)).split(' ')[0]
                        rw+=1
                        pages = convert_from_path(path+'/'+file,250)
                        pages[0].save((isExist2+str(name)+'.png'), 'PNG')
                        ImageCrop((isExist2+str(name))+'.png')
                        sheet.cell(row=rw-1,column=31).value=(str(name)+'.png')           
                except:
                    print('move2   ',rw)
                    shutil.copy(path+'/'+file, isExist1)
                    sheet2.cell(row=rw_fail,column=1).value=file     
                    sheet2.cell(row=rw_fail,column=2).value=path.replace(direc,'')
                    rw_fail+=1
                    
                try:
                    (tables1[4].df).to_excel((direc+'/tables/'+'first6'+file+'.xlsx'),index=False)
                    #sheet.cell(row=rw,column=37).value=list(tables1[4].df[0])[0]
                    sheet.cell(row=rw-1,column=37).value=",".join([(tables1[4].df[0][x]+":"+y) for x,y in enumerate(tables1[4].df[1]) if y != ""])
                except Exception as ex:
                    pass



# col=collar_trim(tables1[3].df)


colordf=pd.DataFrame(color_lis, columns=['color'])
colordf.to_excel(('color-before'+'.xlsx'),index=False)
colordf.drop_duplicates(inplace=True)
colordf.to_excel(('color-after'+'.xlsx'),index=False)
for i in range (0,len(colordf.index)):
    sheet3.cell(row=i+3,column=3).value=(str(colordf.iloc[i,0]).split(':'))[0]
    sheet3.cell(row=i+3,column=2).value=(str(colordf.iloc[i,0]).split(':'))[1]
wb.save(('results.xlsx'))   
print(datetime.now() - startTime) 
wb=openpyxl.load_workbook('results.xlsx')
sheet3=wb.worksheets[2]
df3=pd.read_excel('results.xlsx', usecols=[14,15,19,20,24,25,31,32,33,34,35])

df4=pd.concat([(df3.iloc[:,0]+'+'+df3.iloc[:,1]),(df3.iloc[:,2]+'+'+df3.iloc[:,3]),(df3.iloc[:,4]+'+'+df3.iloc[:,5])])
df4.drop_duplicates(inplace=True)
df4=df4[~df4.str.contains("Recycled",na=False)]
df4=df4[~df4.str.contains("Polyester",na=False)]
df1 = df4.str.split('+',expand=True)
for i in range (2,len(df1.index)):
    sheet3.cell(row=i+1,column=14).value=df1.iloc[i,1]
    sheet3.cell(row=i+1,column=15).value=df1.iloc[i,0]
wb.save(('results2.xlsx'))
# df1.to_excel(('first'+'.xlsx'),index=False)
# df3=pd.read_excel('results.xlsx', usecols=[14,15,19,20,24,25,31,32,33,34,35])
w=[]
df4=pd.concat([df3.iloc[:,9]])
for q in  df4:
   w=w+ str(q).split(',')
collar=pd.DataFrame(w, columns=['collar'])
collar['collar'].replace('nan', np.nan, inplace=True)
collar.dropna(subset=['collar'], inplace=True)
# collar['collar']=collar['collar'][~collar['collar'].str.contains('nan')]
collar.drop_duplicates(inplace=True)
# collar['collar'].to_excel(('first2'+'.xlsx'),index=False)

for q in range (1,len(collar.index)):
   if len((str(collar.iloc[q,0]).split())) > 0: 
    collar.iloc[q,0]=(str(collar.iloc[q,0]).split())[0]


collar.drop_duplicates(inplace=True) 
# collar['collar'].to_excel(('first2'+'.xlsx'),index=False)
for q in range (1,len(collar.index)):
   if len((str(collar.iloc[q,0]).split())) > 0: 
    sheet3.cell(row=q+2,column=6).value=(str(collar.iloc[q,0]).split())[0]
# df4.to_excel(('first2'+'.xlsx'),index=False)   
wb.save(('results2.xlsx'))
w=[]
df4=pd.concat([df3.iloc[:,10]])
for q in  df4:
   w=w+ str(q).split(',')
collar=pd.DataFrame(w, columns=['collar'])
collar['collar'].replace('nan', np.nan, inplace=True)
collar.dropna(subset=['collar'], inplace=True)
# collar['collar']=collar['collar'][~collar['collar'].str.contains('nan')]
collar.drop_duplicates(inplace=True)
for q in range (1,len(collar.index)):
    
   collar.iloc[q,0]=(str(collar.iloc[q,0]).split())[0]
collar.drop_duplicates(inplace=True) 
# collar['collar'].to_excel(('first3'+'.xlsx'),index=False)
for q in range (1,len(collar.index)):
    
    sheet3.cell(row=q+2,column=10).value=(str(collar.iloc[q,0]).split())[0]
# df4.to_excel(('first2'+'.xlsx'),index=False)   
wb.save(('results2.xlsx'))
df3=pd.read_excel('results2.xlsx',sheet_name='Lists', names=[16,17,18])
df4=pd.read_excel('results2.xlsx',sheet_name='Lists', names=[20,21,22])
wb=openpyxl.load_workbook('results2.xlsx')
sheet=wb.worksheets[0]
sheet2=wb.worksheets[1]
sheet3=wb.worksheets[2]
print(df3)
print(df4)
# color_frmt =re.compile(r"[C]\-[0-9]{3}")
collar_lis=[]
color_fin=[]
trim_lis=[]
colors=['EMERALD','GUNMETAL','NAVY','BLACK','WHITE','GOLD','ROYAL','INDIGO','NEW FOREST','PURPLE','MALIBU','CREAM','MAROON','INK',"SCARLET",'GREY','JADE','BROWN','TEAL','MARS RED','REFLEX BLUE','POWDER BLUE','CHARCOAL','NEW GOLD','PINK','OCHRE','BEEWAH BLUE']
j=2
k=2
l=2
for i in range(2,5000):
    if (len(re.findall(re.compile(r"[C]\-[0-9]{3}"),str(sheet3.cell(row=i,column=2).value)))) ==1 and (sheet3.cell(row=i,column=3).value in colors) and  (sheet3.cell(row=i,column=3).value not in color_fin):
        color_fin.append(sheet3.cell(row=i,column=2).value)
        sheet3.cell(row=j,column=2).value=sheet3.cell(row=i,column=2).value
        sheet3.cell(row=j,column=3).value=sheet3.cell(row=i,column=3).value
        j=j+1
        sheet3.cell(row=i,column=2).value=''
        sheet3.cell(row=i,column=3).value=''
    else:
        sheet3.cell(row=i,column=2).value=''
        sheet3.cell(row=i,column=3).value=''
    
    if ((str(sheet3.cell(row=i,column=17).value).lower().startswith('ds')) or (str(sheet3.cell(row=i,column=17).value).lower().startswith('0'))) and ((str(sheet3.cell(row=i,column=17).value)+','+str(sheet3.cell(row=i,column=18).value)+','+str(sheet3.cell(row=i,column=19).value)) not in collar_lis):
        collar_lis.append((str(sheet3.cell(row=i,column=17).value)+','+str(sheet3.cell(row=i,column=18).value)+','+str(sheet3.cell(row=i,column=19).value)))
        sheet3.cell(row=k,column=6).value=sheet3.cell(row=i,column=17).value
        sheet3.cell(row=k,column=7).value=sheet3.cell(row=i,column=18).value
        sheet3.cell(row=k,column=8).value=sheet3.cell(row=i,column=19).value
        k=k+1
        sheet3.cell(row=i,column=17).value=''
        sheet3.cell(row=i,column=18).value=''
        sheet3.cell(row=i,column=19).value=''
    else:
        sheet3.cell(row=i,column=17).value=''
        sheet3.cell(row=i,column=18).value=''
        sheet3.cell(row=i,column=19).value=''

    if ((str(sheet3.cell(row=i,column=21).value).lower().startswith('ds')) or (str(sheet3.cell(row=i,column=21).value).lower().startswith('pt'))) and ((str(sheet3.cell(row=i,column=21).value)+','+str(sheet3.cell(row=i,column=22).value)+','+str(sheet3.cell(row=i,column=23).value)) not in trim_lis):
       
        trim_lis.append((str(sheet3.cell(row=i,column=21).value)+','+str(sheet3.cell(row=i,column=22).value)+','+str(sheet3.cell(row=i,column=23).value)))
        sheet3.cell(row=l,column=10).value=sheet3.cell(row=i,column=21).value
        sheet3.cell(row=l,column=11).value=sheet3.cell(row=i,column=22).value
        sheet3.cell(row=l,column=12).value=sheet3.cell(row=i,column=32).value
        l=l+1
        sheet3.cell(row=i,column=21).value=''
        sheet3.cell(row=i,column=22).value=''
        sheet3.cell(row=i,column=23).value=''
    else:
        sheet3.cell(row=i,column=21).value=''
        sheet3.cell(row=i,column=22).value=''
        sheet3.cell(row=i,column=23).value=''
    if str(sheet.cell(row=i+2,column=8).value).startswith('L  ') or str(sheet.cell(row=i+2,column=8).value).startswith('L '):
        sheet.cell(row=i+2,column=7).value=str(sheet.cell(row=i+2,column=7).value)+'L'
        sheet.cell(row=i+2,column=8).value=str(sheet.cell(row=i+2,column=8).value).replace('L  ','').replace('L ','')
    if str(sheet.cell(row=i+2,column=8).value).startswith('2XS - '):
        sheet.cell(row=i+2,column=7).value=(str(sheet.cell(row=i+2,column=8).value).split('  '))[0]
        sheet.cell(row=i+2,column=8).value=(str(sheet.cell(row=i+2,column=8).value).split('  '))[1]
    if ('2XS - 7X'+'\n'+' MENSL') in str(sheet.cell(row=i+2,column=7).value):
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if ('2XS - 7X '+'\n'+' MENSL') in str(sheet.cell(row=i+2,column=7).value):
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if ('2XS - 7X L') in str(sheet.cell(row=i+2,column=7).value):
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if ('2XS - 7XL'+'\n '+'\n'+'MENS') in str(sheet.cell(row=i+2,column=7).value):
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if ('2XS - 7XL'+'\n'+' MENS') in str(sheet.cell(row=i+2,column=7).value):
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if ('2XS - 7XL') in str(sheet.cell(row=i+2,column=7).value) or ('2XS- 7XL' in str(sheet.cell(row=i+2,column=7).value))  :
        (sheet.cell(row=i+2,column=7).value)='2XS - 7XL'
    if (  str(sheet.cell(row=i+2,column=7).value)).startswith('XS - 7XL'):
        (sheet.cell(row=i+2,column=7).value)='XS - 7XL'
    if   ('XS - 5XL' in str(sheet.cell(row=i+2,column=7).value)):
        (sheet.cell(row=i+2,column=7).value)='XS - 5XL'
    if '  MENS' in str(sheet.cell(row=i+2,column=6).value) or ' MENS' in str(sheet.cell(row=i+2,column=6).value):
        (sheet.cell(row=i+2,column=6).value)=(str(sheet.cell(row=i+2,column=6).value)).replace('  MENS','').replace(' MENS','')
    
    if (  str(sheet.cell(row=i+2,column=11).value)).startswith('6 - 22') or (  str(sheet.cell(row=i+2,column=11).value)).startswith('6-22'):
        sheet.cell(row=i+2,column=10).value='6 - 22'
        sheet.cell(row=i+2,column=11).value='DSS'+(str(sheet.cell(row=i+2,column=11).value).split('DSS')[1])
    if str(sheet.cell(row=i+2,column=9).value)=='6 - 22 ':
        sheet.cell(row=i+2,column=10).value='6 - 22'
        sheet.cell(row=i+2,column=9).value=''
    if ('LADIES ' in str(sheet.cell(row=i+2,column=9).value)) or ' LADIES' in str(sheet.cell(row=i+2,column=9).value):
        (sheet.cell(row=i+2,column=9).value)=(str(sheet.cell(row=i+2,column=9).value)).replace('LADIES ','').replace(' LADIES','').replace('\n','')
    if ('LADIES ' in str(sheet.cell(row=i+2,column=10).value)) or ' LADIES' in str(sheet.cell(row=i+2,column=10).value):
        (sheet.cell(row=i+2,column=10).value)=(str(sheet.cell(row=i+2,column=10).value)).replace('LADIES ','').replace(' LADIES','').replace('\n','')
    if (' JUNIORS' in str(sheet.cell(row=i+2,column=12).value)) or '   JUNIORS' in str(sheet.cell(row=i+2,column=12).value) or '  JUNIORS' in str(sheet.cell(row=i+2,column=12).value) or ' JUNIORS' in str(sheet.cell(row=i+2,column=12).value) or str(sheet.cell(row=i+2,column=12).value).startswith('JUNIORS'):
        (sheet.cell(row=i+2,column=12).value)=(str(sheet.cell(row=i+2,column=12).value)).replace('   JUNIORS','').replace('  JUNIORS','').replace(' JUNIORS','').replace('JUNIORS','').replace('\n','')
    if ('J4 - J14' in str(sheet.cell(row=i+2,column=12).value)) and ('DSSIJPT211E' in str(sheet.cell(row=i+2,column=13).value)):
        sheet.cell(row=i+2,column=12).value='DSSIJPT211E'
        sheet.cell(row=i+2,column=13).value='J4 - J14'
    if ('J4 - J14' in str(sheet.cell(row=i+2,column=12).value)) and ('DSSIJPT211E' in str(sheet.cell(row=i+2,column=12).value)) and  ('DSSPEC-7009A' in str(sheet.cell(row=i+2,column=13).value)):
        sheet.cell(row=i+2,column=12).value='DSSIJPT211E'
        sheet.cell(row=i+2,column=13).value='J4 - J14'
        sheet.cell(row=i+2,column=14).value='DSSPEC-7009A'
    if (str(sheet.cell(row=i+2,column=13).value).startswith('JUNIORS  ')) or (str(sheet.cell(row=i+2,column=13).value).startswith('JUNIORS ')):
        sheet.cell(row=i+2,column=13).value=str(sheet.cell(row=i+2,column=13).value).replace('JUNIORS  ','').replace('JUNIORS ','')
    if (str(sheet.cell(row=i+2,column=14).value).startswith('J4 - J14')) or (str(sheet.cell(row=i+2,column=14).value).startswith('J4-J14')):
        sheet.cell(row=i+2,column=13).value=str(sheet.cell(row=i+2,column=14).value).split('\n')[0]
        sheet.cell(row=i+2,column=14).value=str(sheet.cell(row=i+2,column=14).value).split('\n')[1]
    if (sheet.cell(row=i+2,column=6).value) != None:
        (sheet.cell(row=i+2,column=6).value)=(str(sheet.cell(row=i+2,column=6).value)).replace(' ','')
    if (sheet.cell(row=i+2,column=8).value) != None:
        (sheet.cell(row=i+2,column=8).value)=(str(sheet.cell(row=i+2,column=8).value)).replace(' ','')
    if (sheet.cell(row=i+2,column=9).value) != None:
        (sheet.cell(row=i+2,column=9).value)=(str(sheet.cell(row=i+2,column=9).value)).replace(' ','')
    if (sheet.cell(row=i+2,column=11).value) != None:
        (sheet.cell(row=i+2,column=11).value)=(str(sheet.cell(row=i+2,column=11).value)).replace(' ','')
    if (sheet.cell(row=i+2,column=12).value) != None:
        (sheet.cell(row=i+2,column=12).value)=(str(sheet.cell(row=i+2,column=12).value)).replace(' ','')
    if (sheet.cell(row=i+2,column=14).value) != None:
        (sheet.cell(row=i+2,column=14).value)=(str(sheet.cell(row=i+2,column=14).value)).replace(' ','')
    if (' JUNIOR' in str(sheet.cell(row=i+2,column=12).value)) or '   JUNIOR' in str(sheet.cell(row=i+2,column=12).value) or 'JUNIOR' in str(sheet.cell(row=i+2,column=12).value) or '  JUNIOR' in str(sheet.cell(row=i+2,column=12).value) or ' JUNIOR' in str(sheet.cell(row=i+2,column=12).value) or str(sheet.cell(row=i+2,column=12).value).startswith('JUNIOR'):
        (sheet.cell(row=i+2,column=12).value)=(str(sheet.cell(row=i+2,column=12).value)).replace('   JUNIOR','').replace('  JUNIOR','').replace(' JUNIOR','').replace('JUNIOR','').replace('\n','')
wb.save(('results_final.xlsx'))

# df3.drop_duplicates(subset=[df3.columns[1], df3.columns[2],df3.columns[0]],inplace=True)
# df4.drop_duplicates(subset=[df4.columns[1], df4.columns[2],df4.columns[0]],inplace=True)
# df3.to_excel(('collar'+'.xlsx'),index=False)
# df4.to_excel(('trim'+'.xlsx'),index=False)

